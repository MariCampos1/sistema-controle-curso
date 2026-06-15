from flask import Flask, render_template, request, redirect, send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
from datetime import datetime
from openpyxl import Workbook
from io import BytesIO
from io import BytesIO
from urllib.parse import urlencode
import os
import uuid
import base64
import resend


app = Flask(__name__)


# --------------------------------------------------
# CONFIGURAÇÃO DO BANCO
# --------------------------------------------------

database_url = os.environ.get("DATABASE_URL")

# Banco local para testes no computador
if not database_url:
    database_url = "sqlite:///cadastro_local.db"

if database_url.startswith("postgresql://"):
    database_url = database_url.replace(
        "postgresql://",
        "postgresql+psycopg2://",
        1
    )

app.config["SQLALCHEMY_DATABASE_URI"] = database_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_pre_ping": True
}

db = SQLAlchemy(app)


# --------------------------------------------------
# MODELO: ESTUDANTE
# --------------------------------------------------

class Estudante(db.Model):
    __tablename__ = "estudantes"

    id = db.Column(
        db.String(36),
        primary_key=True,
        default=lambda: str(uuid.uuid4())
    )

    nome = db.Column(
        db.String(150),
        nullable=False
    )

    criado_em = db.Column(
        db.DateTime(timezone=True),
        server_default=db.func.now()
    )

    matriculas = db.relationship(
        "Matricula",
        back_populates="estudante",
        cascade="all, delete-orphan"
    )

class Curso(db.Model):
    __tablename__ = "cursos"

    id = db.Column(
        db.String(36),
        primary_key=True,
        default=lambda: str(uuid.uuid4())
    )

    nome = db.Column(
        db.String(150),
        nullable=False
    )

    professor = db.Column(
        db.String(150),
        nullable=False
    )

    criado_em = db.Column(
        db.DateTime(timezone=True),
        server_default=db.func.now()
    )
    
# --------------------------------------------------
# MODELO: MATRÍCULA
# --------------------------------------------------

class Matricula(db.Model):
    __tablename__ = "matriculas"

    id = db.Column(
        db.String(36),
        primary_key=True,
        default=lambda: str(uuid.uuid4())
    )

    estudante_id = db.Column(
        db.String(36),
        db.ForeignKey(
            "estudantes.id",
            ondelete="CASCADE"
        ),
        nullable=False
    )

    curso = db.Column(
        db.String(150),
        nullable=False
    )

    professor = db.Column(
        db.String(150),
        nullable=False
    )

    presente = db.Column(
        db.Boolean,
        nullable=False,
        default=False
    )

    data_presenca = db.Column(
        db.String(10),
        nullable=False,
        default=""
    )

    pagamento = db.Column(
        db.Boolean,
        nullable=False,
        default=False
    )

    alimento = db.Column(
        db.Boolean,
        nullable=False,
        default=False
    )

    estudante = db.relationship(
        "Estudante",
        back_populates="matriculas"
    )

    __table_args__ = (
        db.UniqueConstraint(
            "estudante_id",
            "curso",
            name="matricula_unica_por_curso"
        ),
    )

    # Permite que o HTML continue usando a.nome
    @property
    def nome(self):
        return self.estudante.nome


with app.app_context():
    db.create_all()


# --------------------------------------------------
# TELA PRINCIPAL
# --------------------------------------------------

@app.route("/")
def painel():
    return render_template("painel.html")

@app.route("/alunos")
def index():
    curso_filtro = request.args.get(
        "curso",
        ""
    ).strip()

    professor_filtro = request.args.get(
        "professor",
        ""
    ).strip()

    busca = request.args.get(
        "busca",
        ""
    ).strip()

# Cursos usados no filtro da tabela
    cursos_resultado = (
        db.session.query(Matricula.curso)
        .filter(Matricula.curso != "")
        .distinct()
        .order_by(Matricula.curso)
        .all()
    )

    cursos = [
        resultado[0]
        for resultado in cursos_resultado
    ]

# Cursos disponíveis para novas matrículas
    cursos_cadastro = (
        db.session.query(Curso)
        .order_by(Curso.nome.asc())
        .all()
    )

    professores_resultado = (
        db.session.query(Matricula.professor)
        .filter(Matricula.professor != "")
        .distinct()
        .order_by(Matricula.professor)
        .all()
    )

    professores = [
        resultado[0]
        for resultado in professores_resultado
    ]

    consulta = (
        db.session.query(Matricula)
        .join(Estudante)
    )

    if curso_filtro:
        consulta = consulta.filter(
            Matricula.curso == curso_filtro
        )

    if professor_filtro:
        consulta = consulta.filter(
            Matricula.professor == professor_filtro
        )

    if busca:
        consulta = consulta.filter(
            Estudante.nome.ilike(f"%{busca}%")
        )

    matriculas = consulta.order_by(
        Estudante.nome.asc(),
        Matricula.curso.asc()
    ).all()

    # Conta pessoas sem duplicar quem possui dois cursos
    total_alunos = len({
        matricula.estudante_id
        for matricula in matriculas
    })

    total_presentes = sum(
        1 for matricula in matriculas
        if matricula.presente
    )

    total_pagamentos = sum(
        1 for matricula in matriculas
        if matricula.pagamento
    )

    total_alimentos = sum(
        1 for matricula in matriculas
        if matricula.alimento
    )

    return render_template(
        "index.html",

        # Mantém o nome "alunos" para não quebrar seu HTML
        alunos=matriculas,

        cursos=cursos,
        cursos_cadastro=cursos_cadastro,
        curso_filtro=curso_filtro,
        professores=professores,
        professor_filtro=professor_filtro,
        busca=busca,
        total_alunos=total_alunos,
        total_presentes=total_presentes,
        total_pagamentos=total_pagamentos,
        total_alimentos=total_alimentos
    )


# --------------------------------------------------
# ADICIONAR MATRÍCULA
# --------------------------------------------------

@app.route("/adicionar", methods=["POST"])
def adicionar():
    nome = request.form.get(
        "nome",
        ""
    ).strip()

    curso_id = request.form.get(
        "curso_id",
        ""
    ).strip()

    if not nome or not curso_id:
        return redirect("/alunos")

    curso_escolhido = db.session.get(
        Curso,
        curso_id
    )

    if curso_escolhido is None:
        return "Curso não encontrado", 404

    curso = curso_escolhido.nome
    professor = curso_escolhido.professor

    estudante = (
        db.session.query(Estudante)
        .filter(
            func.lower(Estudante.nome)
            == nome.lower()
        )
        .first()
    )

    if estudante is None:
        estudante = Estudante(nome=nome)
        db.session.add(estudante)
        db.session.flush()

    matricula_existente = (
        db.session.query(Matricula)
        .filter(
            Matricula.estudante_id == estudante.id,
            func.lower(Matricula.curso)
            == curso.lower()
        )
        .first()
    )

    if matricula_existente:
        return (
            "Este aluno já está matriculado nesse curso. "
            '<a href="/">Voltar</a>',
            400
        )

    nova_matricula = Matricula(
        estudante_id=estudante.id,
        curso=curso,
        professor=professor,
        presente=False,
        data_presenca="",
        pagamento=False,
        alimento=False
    )

    db.session.add(nova_matricula)

    try:
        db.session.commit()

    except IntegrityError:
        db.session.rollback()

        return (
            "Não foi possível criar a matrícula. "
            '<a href="/">Voltar</a>',
            400
        )

    return redirect("/alunos")

    # Procura um estudante já cadastrado com esse nome
    estudante = (
        db.session.query(Estudante)
        .filter(
            func.lower(Estudante.nome)
            == nome.lower()
        )
        .first()
    )

    # Se não existir, cria o estudante
    if estudante is None:
        estudante = Estudante(nome=nome)
        db.session.add(estudante)
        db.session.flush()

    # Verifica se já existe matrícula nesse curso
    matricula_existente = (
        db.session.query(Matricula)
        .filter(
            Matricula.estudante_id == estudante.id,
            func.lower(Matricula.curso)
            == curso.lower()
        )
        .first()
    )

    if matricula_existente:
        return (
            "Este aluno já está matriculado nesse curso. "
            '<a href="/">Voltar</a>',
            400
        )

    nova_matricula = Matricula(
        estudante_id=estudante.id,
        curso=curso,
        professor=professor,
        presente=False,
        data_presenca="",
        pagamento=False,
        alimento=False
    )

    db.session.add(nova_matricula)

    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()

        return (
            "Não foi possível criar a matrícula. "
            "Verifique se ela já existe. "
            '<a href="/">Voltar</a>',
            400
        )

    return redirect("/alunos")


# --------------------------------------------------
# PRESENÇA POR CURSO
# --------------------------------------------------

@app.route("/presenca/<id_matricula>")
def presenca(id_matricula):
    matricula = db.session.get(
        Matricula,
        id_matricula
    )

    if matricula is None:
        return "Matrícula não encontrada", 404

    if matricula.presente:
        matricula.presente = False
        matricula.data_presenca = ""
    else:
        matricula.presente = True
        matricula.data_presenca = (
            datetime.now().strftime("%d/%m/%Y")
        )

    db.session.commit()

    return redirect("/alunos")


# --------------------------------------------------
# PAGAMENTO POR CURSO
# --------------------------------------------------

@app.route("/pagamento/<id_matricula>")
def pagamento(id_matricula):
    matricula = db.session.get(
        Matricula,
        id_matricula
    )

    if matricula is None:
        return "Matrícula não encontrada", 404

    matricula.pagamento = not matricula.pagamento

    db.session.commit()

    return redirect("/alunos")


# --------------------------------------------------
# ALIMENTO POR CURSO
# --------------------------------------------------

@app.route("/alimento/<id_matricula>")
def alimento(id_matricula):
    matricula = db.session.get(
        Matricula,
        id_matricula
    )

    if matricula is None:
        return "Matrícula não encontrada", 404

    matricula.alimento = not matricula.alimento

    db.session.commit()

    return redirect("/alunos")


# --------------------------------------------------
# EXCLUIR MATRÍCULA
# --------------------------------------------------

@app.route("/excluir/<id_matricula>")
def excluir(id_matricula):
    matricula = db.session.get(
        Matricula,
        id_matricula
    )

    if matricula is None:
        return "Matrícula não encontrada", 404

    estudante = matricula.estudante

    db.session.delete(matricula)
    db.session.flush()

    # Se o estudante ficou sem nenhum curso,
    # remove também seu cadastro principal
    quantidade_matriculas = (
        db.session.query(Matricula)
        .filter(
            Matricula.estudante_id == estudante.id
        )
        .count()
    )

    if quantidade_matriculas == 0:
        db.session.delete(estudante)

    db.session.commit()

    return redirect("/alunos")


# --------------------------------------------------
# EDITAR PROFESSOR DA MATRÍCULA
# --------------------------------------------------

@app.route(
    "/editar_professor/<id_matricula>",
    methods=["GET", "POST"]
)
def editar_professor(id_matricula):
    matricula = db.session.get(
        Matricula,
        id_matricula
    )

    if matricula is None:
        return "Matrícula não encontrada", 404

    if request.method == "POST":
        novo_professor = request.form.get(
            "professor",
            ""
        ).strip()

        if novo_professor:
            matricula.professor = novo_professor
            db.session.commit()

        return redirect("/alunos")

    return render_template(
        "editar_professor.html",
        aluno=matricula
    )


# --------------------------------------------------
# NOVA CHAMADA
# --------------------------------------------------

@app.route("/nova_chamada")
def nova_chamada():
    matriculas = db.session.query(
        Matricula
    ).all()

    for matricula in matriculas:
        matricula.presente = False
        matricula.data_presenca = ""

    db.session.commit()

    return redirect("/alunos")


# --------------------------------------------------
# EXPORTAR
# --------------------------------------------------

@app.route("/exportar")
def exportar():
    matriculas = (
        db.session.query(Matricula)
        .join(Estudante)
        .order_by(
            Estudante.nome.asc(),
            Matricula.curso.asc()
        )
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Matrículas"

    ws.append([
        "Aluno",
        "Curso",
        "Professor",
        "Presença",
        "Data da Presença",
        "Pagamento",
        "Alimento"
    ])

    for matricula in matriculas:
        ws.append([
            matricula.estudante.nome,
            matricula.curso,
            matricula.professor,
            (
                "Presente"
                if matricula.presente
                else "Ausente"
            ),
            matricula.data_presenca,
            (
                "Pago"
                if matricula.pagamento
                else "Pendente"
            ),
            (
                "Entregue"
                if matricula.alimento
                else "Pendente"
            )
        ])

    arquivo = BytesIO()

    wb.save(arquivo)
    arquivo.seek(0)

    return send_file(
        arquivo,
        as_attachment=True,
        download_name="relatorio_matriculas.xlsx",
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

# --------------------------------------------------
# GERENCIAR CURSOS
# --------------------------------------------------

@app.route("/cursos")
def gerenciar_cursos():
    cursos = (
        db.session.query(Curso)
        .order_by(Curso.nome.asc())
        .all()
    )

    return render_template(
        "cursos.html",
        cursos=cursos
    )


# --------------------------------------------------
# ADICIONAR CURSO
# --------------------------------------------------

@app.route("/cursos/adicionar", methods=["POST"])
def adicionar_curso():
    nome = request.form.get("nome", "").strip()
    professor = request.form.get("professor", "").strip()

    if not nome or not professor:
        return redirect("/cursos")

    curso_existente = (
        db.session.query(Curso)
        .filter(
            func.lower(Curso.nome) == nome.lower()
        )
        .first()
    )

    if curso_existente:
        return (
            "Esse curso já está cadastrado. "
            '<a href="/cursos">Voltar</a>',
            400
        )

    novo_curso = Curso(
        nome=nome,
        professor=professor
    )

    db.session.add(novo_curso)
    db.session.commit()

    return redirect("/cursos")


# --------------------------------------------------
# EDITAR CURSO E PROFESSOR
# --------------------------------------------------

@app.route("/cursos/editar/<id_curso>", methods=["POST"])
def editar_curso(id_curso):
    curso = db.session.get(Curso, id_curso)

    if curso is None:
        return "Curso não encontrado", 404

    novo_nome = request.form.get("nome", "").strip()
    novo_professor = request.form.get(
        "professor",
        ""
    ).strip()

    if not novo_nome or not novo_professor:
        return redirect("/cursos")

    curso_duplicado = (
        db.session.query(Curso)
        .filter(
            func.lower(Curso.nome) == novo_nome.lower(),
            Curso.id != id_curso
        )
        .first()
    )

    if curso_duplicado:
        return (
            "Já existe outro curso com esse nome. "
            '<a href="/cursos">Voltar</a>',
            400
        )

    nome_antigo = curso.nome

    # Atualiza também as matrículas já existentes.
    matriculas = (
        db.session.query(Matricula)
        .filter(
            func.lower(Matricula.curso)
            == nome_antigo.lower()
        )
        .all()
    )

    for matricula in matriculas:
        matricula.curso = novo_nome
        matricula.professor = novo_professor

    curso.nome = novo_nome
    curso.professor = novo_professor

    db.session.commit()

    return redirect("/cursos")


# --------------------------------------------------
# EXCLUIR CURSO
# --------------------------------------------------

@app.route("/cursos/excluir/<id_curso>", methods=["POST"])
def excluir_curso(id_curso):
    curso = db.session.get(Curso, id_curso)

    if curso is None:
        return "Curso não encontrado", 404

    matriculas_existentes = (
        db.session.query(Matricula)
        .filter(
            func.lower(Matricula.curso)
            == curso.nome.lower()
        )
        .count()
    )

    if matriculas_existentes > 0:
        return (
            "Esse curso não pode ser excluído porque possui "
            "alunos matriculados. "
            '<a href="/cursos">Voltar</a>',
            400
        )

    db.session.delete(curso)
    db.session.commit()

    return redirect("/cursos")

@app.route("/chamada")
def pagina_chamada():
    curso_selecionado = request.args.get(
        "curso",
        ""
    ).strip()

    cursos = (
        db.session.query(Curso)
        .order_by(Curso.nome.asc())
        .all()
    )

    matriculas = []

    if curso_selecionado:
        matriculas = (
            db.session.query(Matricula)
            .join(Estudante)
            .filter(
                func.lower(Matricula.curso)
                == curso_selecionado.lower()
            )
            .order_by(Estudante.nome.asc())
            .all()
        )

    total_presentes = sum(
        1 for matricula in matriculas
        if matricula.presente
    )

    return render_template(
        "chamada.html",
        cursos=cursos,
        curso_selecionado=curso_selecionado,
        matriculas=matriculas,
        total_presentes=total_presentes
    )

@app.route("/controles")
def pagina_controles():
    curso_selecionado = request.args.get(
        "curso",
        ""
    ).strip()

    cursos = (
        db.session.query(Curso)
        .order_by(Curso.nome.asc())
        .all()
    )

    matriculas = []

    if curso_selecionado:
        matriculas = (
            db.session.query(Matricula)
            .join(Estudante)
            .filter(
                func.lower(Matricula.curso)
                == curso_selecionado.lower()
            )
            .order_by(Estudante.nome.asc())
            .all()
        )

    total_pagamentos = sum(
        1 for matricula in matriculas
        if matricula.pagamento
    )

    total_alimentos = sum(
        1 for matricula in matriculas
        if matricula.alimento
    )

    return render_template(
        "controles.html",
        cursos=cursos,
        curso_selecionado=curso_selecionado,
        matriculas=matriculas,
        total_pagamentos=total_pagamentos,
        total_alimentos=total_alimentos
    )

@app.route("/relatorios")
def pagina_relatorios():

    enviado = request.args.get("enviado") == "1"
    erro_envio = request.args.get("erro_envio") == "1"

    curso_filtro = request.args.get("curso", "").strip()
    professor_filtro = request.args.get("professor", "").strip()
    busca = request.args.get("busca", "").strip()

    cursos = (
        db.session.query(Curso)
        .order_by(Curso.nome.asc())
        .all()
    )

    professores_resultado = (
        db.session.query(Matricula.professor)
        .filter(Matricula.professor != "")
        .distinct()
        .order_by(Matricula.professor.asc())
        .all()
    )

    professores = [
        resultado[0]
        for resultado in professores_resultado
    ]

    consulta = (
        db.session.query(Matricula)
        .join(Estudante)
    )

    if curso_filtro:
        consulta = consulta.filter(
            func.lower(Matricula.curso)
            == curso_filtro.lower()
        )

    if professor_filtro:
        consulta = consulta.filter(
            func.lower(Matricula.professor)
            == professor_filtro.lower()
        )

    if busca:
        consulta = consulta.filter(
            Estudante.nome.ilike(f"%{busca}%")
        )

    matriculas = consulta.order_by(
        Estudante.nome.asc(),
        Matricula.curso.asc()
    ).all()

    total_alunos = len({
        matricula.estudante_id
        for matricula in matriculas
    })

    total_matriculas = len(matriculas)

    total_presentes = sum(
        1 for matricula in matriculas
        if matricula.presente
    )

    total_pagamentos = sum(
        1 for matricula in matriculas
        if matricula.pagamento
    )

    total_alimentos = sum(
        1 for matricula in matriculas
        if matricula.alimento
    )

    return render_template(
        "relatorios.html",
        cursos=cursos,
        professores=professores,
        matriculas=matriculas,
        curso_filtro=curso_filtro,
        professor_filtro=professor_filtro,
        busca=busca,
        total_alunos=total_alunos,
        total_matriculas=total_matriculas,
        total_presentes=total_presentes,
        total_pagamentos=total_pagamentos,
        total_alimentos=total_alimentos,
        enviado=enviado,
        erro_envio=erro_envio
    )

@app.route("/chamada/presenca/<id_matricula>")
def chamada_presenca(id_matricula):
    matricula = db.session.get(
        Matricula,
        id_matricula
    )

    if matricula is None:
        return "Matrícula não encontrada", 404

    if matricula.presente:
        matricula.presente = False
        matricula.data_presenca = ""
    else:
        matricula.presente = True
        matricula.data_presenca = (
            datetime.now().strftime("%d/%m/%Y")
        )

    db.session.commit()

    return redirect(
        f"/chamada?curso={matricula.curso}"
    )

@app.route("/controles/pagamento/<id_matricula>")
def controles_pagamento(id_matricula):
    matricula = db.session.get(
        Matricula,
        id_matricula
    )

    if matricula is None:
        return "Matrícula não encontrada", 404

    matricula.pagamento = not matricula.pagamento

    db.session.commit()

    return redirect(
        f"/controles?curso={matricula.curso}"
    )

@app.route("/controles/alimento/<id_matricula>")
def controles_alimento(id_matricula):
    matricula = db.session.get(
        Matricula,
        id_matricula
    )

    if matricula is None:
        return "Matrícula não encontrada", 404

    matricula.alimento = not matricula.alimento

    db.session.commit()

    return redirect(
        f"/controles?curso={matricula.curso}"
    )

@app.route("/relatorios/exportar")
def exportar_relatorio():
    curso_filtro = request.args.get("curso", "").strip()
    professor_filtro = request.args.get("professor", "").strip()
    busca = request.args.get("busca", "").strip()

    consulta = (
        db.session.query(Matricula)
        .join(Estudante)
    )

    if curso_filtro:
        consulta = consulta.filter(
            func.lower(Matricula.curso)
            == curso_filtro.lower()
        )

    if professor_filtro:
        consulta = consulta.filter(
            func.lower(Matricula.professor)
            == professor_filtro.lower()
        )

    if busca:
        consulta = consulta.filter(
            Estudante.nome.ilike(f"%{busca}%")
        )

    matriculas = consulta.order_by(
        Estudante.nome.asc(),
        Matricula.curso.asc()
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    ws.append([
        "Aluno",
        "Curso",
        "Professor",
        "Presença",
        "Data da Presença",
        "Pagamento",
        "Alimento"
    ])

    for matricula in matriculas:
        ws.append([
            matricula.estudante.nome,
            matricula.curso,
            matricula.professor,
            "Presente" if matricula.presente else "Ausente",
            matricula.data_presenca,
            "Pago" if matricula.pagamento else "Pendente",
            "Entregue" if matricula.alimento else "Pendente"
        ])

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    return send_file(
        arquivo,
        as_attachment=True,
        download_name="relatorio_filtrado.xlsx",
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

@app.route("/relatorios/enviar")
def enviar_relatorio_email():
    curso_filtro = request.args.get(
        "curso",
        ""
    ).strip()

    professor_filtro = request.args.get(
        "professor",
        ""
    ).strip()

    busca = request.args.get(
        "busca",
        ""
    ).strip()

    consulta = (
        db.session.query(Matricula)
        .join(Estudante)
    )

    if curso_filtro:
        consulta = consulta.filter(
            func.lower(Matricula.curso)
            == curso_filtro.lower()
        )

    if professor_filtro:
        consulta = consulta.filter(
            func.lower(Matricula.professor)
            == professor_filtro.lower()
        )

    if busca:
        consulta = consulta.filter(
            Estudante.nome.ilike(f"%{busca}%")
        )

    matriculas = consulta.order_by(
        Estudante.nome.asc(),
        Matricula.curso.asc()
    ).all()

    # Cria o Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    ws.append([
        "Aluno",
        "Curso",
        "Professor",
        "Presença",
        "Data da Presença",
        "Pagamento",
        "Alimento"
    ])

    for matricula in matriculas:
        ws.append([
            matricula.estudante.nome,
            matricula.curso,
            matricula.professor,
            (
                "Presente"
                if matricula.presente
                else "Ausente"
            ),
            matricula.data_presenca,
            (
                "Pago"
                if matricula.pagamento
                else "Pendente"
            ),
            (
                "Entregue"
                if matricula.alimento
                else "Pendente"
            )
        ])

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    arquivo_base64 = base64.b64encode(
        arquivo.getvalue()
    ).decode("utf-8")

    api_key = os.environ.get("RESEND_API_KEY")
    email_secretaria = os.environ.get(
        "EMAIL_SECRETARIA"
    )
    email_remetente = os.environ.get(
        "EMAIL_REMETENTE"
    )

    filtros_retorno = {
        "curso": curso_filtro,
        "professor": professor_filtro,
        "busca": busca
    }

    if (
        not api_key
        or not email_secretaria
        or not email_remetente
    ):
        filtros_retorno["erro_envio"] = "1"

        return redirect(
            "/relatorios?"
            + urlencode(filtros_retorno)
        )

    resend.api_key = api_key

    descricao_filtro = curso_filtro or "Todos os cursos"

    try:
        params: resend.Emails.SendParams = {
            "from": email_remetente,
            "to": [email_secretaria],
            "subject": (
                f"Relatório de alunos — "
                f"{descricao_filtro}"
            ),
            "html": (
                "<h2>Relatório do sistema de cursos</h2>"
                f"<p>Curso: {descricao_filtro}</p>"
                f"<p>Total de matrículas: "
                f"{len(matriculas)}</p>"
                "<p>O relatório está anexado "
                "a este e-mail.</p>"
            ),
            "attachments": [
                {
                    "content": arquivo_base64,
                    "filename": "relatorio_alunos.xlsx"
                }
            ]
        }

        resend.Emails.send(params)

        filtros_retorno["enviado"] = "1"

    except Exception as erro:
        print(
            "Erro ao enviar relatório:",
            erro
        )

        filtros_retorno["erro_envio"] = "1"

    return redirect(
        "/relatorios?"
        + urlencode(filtros_retorno)
    )

# --------------------------------------------------
# RODAR SERVIDOR
# --------------------------------------------------

if __name__ == "__main__":
    app.run(debug=True)